#!/usr/bin/env python3
"""
email-attachments.py — Scan all daily emails, extract body text and attachment
content, and append to the daily memory file.

Usage:
    python3 scripts/email-attachments.py [--date YYYY-MM-DD] [--dry-run]

Runs as part of the midnight WITA cron, before employee_context_update_run.py.
"""

import argparse
import base64
import json
import os
import subprocess
import sys
from datetime import datetime, timedelta, timezone

# --- Config ---
GOG_PASSWORD = "openclaw-hok-2026"
ACCOUNT = "ops@houseofkairos.com"
WORKSPACE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MEMORY_DIR = os.path.join(WORKSPACE, "memory")
MAX_TEXT_CHARS = 2000
WITA = timezone(timedelta(hours=8))

EXTRACTABLE_EXTENSIONS = {
    ".pdf", ".docx", ".doc", ".xlsx", ".xls", ".csv", ".txt",
}

IDEMPOTENCY_HEADER = "### Email Content"


def run_gog(args):
    """Run a gog CLI command and return stdout, or None on failure."""
    env = os.environ.copy()
    env["GOG_KEYRING_PASSWORD"] = GOG_PASSWORD
    cmd = ["gog"] + args + ["-a", ACCOUNT]
    result = subprocess.run(cmd, capture_output=True, text=True, env=env)
    if result.returncode != 0:
        print(f"  [ERROR] gog {' '.join(args[:3])}: {result.stderr.strip()}", file=sys.stderr)
        return None
    return result.stdout


def search_emails(date_str):
    """Search for all emails on the given date."""
    dt = datetime.strptime(date_str, "%Y-%m-%d")
    next_day = dt + timedelta(days=1)
    query = f"after:{dt.strftime('%Y/%m/%d')} before:{next_day.strftime('%Y/%m/%d')}"
    output = run_gog(["gmail", "search", query, "-j"])
    if not output or not output.strip():
        return []
    try:
        data = json.loads(output)
    except json.JSONDecodeError:
        print("  [ERROR] Failed to parse search results JSON", file=sys.stderr)
        return []

    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        return data.get("messages", data.get("threads", []))
    return []


def get_message(msg_id):
    """Fetch the full message with MIME parts."""
    output = run_gog(["gmail", "get", msg_id, "--format=full", "-j"])
    if not output:
        return None
    try:
        return json.loads(output)
    except json.JSONDecodeError:
        return None


# --- MIME parsing ---

def decode_body_data(data):
    """Decode base64url-encoded body data from Gmail API."""
    if not data:
        return ""
    try:
        # Gmail uses URL-safe base64 without padding
        padded = data + "=" * (4 - len(data) % 4) if len(data) % 4 else data
        return base64.urlsafe_b64decode(padded).decode("utf-8", errors="replace")
    except Exception:
        return ""


def find_body_text(payload):
    """Extract plain-text body from MIME payload tree. Returns string or empty."""
    mime = payload.get("mimeType", "")
    parts = payload.get("parts", [])

    # Single-part message (text/plain at top level)
    if mime == "text/plain" and not parts:
        data = payload.get("body", {}).get("data", "")
        return decode_body_data(data)

    # Walk parts: prefer text/plain, fall back to text/html stripped
    plain = ""
    html = ""
    for part in parts:
        part_mime = part.get("mimeType", "")
        if part_mime == "text/plain":
            data = part.get("body", {}).get("data", "")
            plain = decode_body_data(data)
        elif part_mime == "text/html" and not html:
            data = part.get("body", {}).get("data", "")
            html = decode_body_data(data)
        elif part_mime.startswith("multipart/"):
            # Recurse into nested multipart
            nested = find_body_text(part)
            if nested:
                plain = plain or nested

    if plain:
        return plain

    # Fallback: strip HTML tags crudely
    if html:
        import re
        text = re.sub(r"<[^>]+>", " ", html)
        text = re.sub(r"\s+", " ", text).strip()
        return text

    return ""


def find_attachments(payload):
    """Recursively find attachment parts in a MIME payload tree."""
    attachments = []
    parts = payload.get("parts", [])

    if not parts:
        if payload.get("filename") and payload.get("body", {}).get("attachmentId"):
            attachments.append(payload)
        return attachments

    for part in parts:
        if part.get("filename") and part.get("body", {}).get("attachmentId"):
            attachments.append(part)
        elif part.get("parts"):
            attachments.extend(find_attachments(part))

    return attachments


def download_attachment(msg_id, att_id, filename):
    """Download an attachment to /tmp and return the file path, or None."""
    safe_name = "".join(c if c.isalnum() or c in ".-_" else "_" for c in filename)
    out_path = f"/tmp/att_{msg_id}_{safe_name}"
    run_gog(["gmail", "attachment", msg_id, att_id, f"--out={out_path}"])
    if os.path.exists(out_path):
        return out_path
    return None


# --- Attachment text extraction ---

def extract_text(filepath, filename):
    """Extract text from a file based on its extension. Returns text or None."""
    ext = os.path.splitext(filename.lower())[1]
    if ext == ".pdf":
        return _extract_pdf(filepath)
    if ext in (".docx", ".doc"):
        return _extract_docx(filepath)
    if ext in (".xlsx", ".xls"):
        return _extract_xlsx(filepath)
    if ext in (".csv", ".txt"):
        return _extract_plaintext(filepath)
    return None


def _extract_pdf(filepath):
    try:
        result = subprocess.run(
            ["pdftotext", filepath, "-"],
            capture_output=True, text=True, timeout=30,
        )
        if result.returncode == 0 and result.stdout.strip():
            return result.stdout.strip()[:MAX_TEXT_CHARS]
    except (subprocess.TimeoutExpired, FileNotFoundError):
        pass
    return None


def _extract_docx(filepath):
    try:
        from docx import Document
        doc = Document(filepath)
        text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        return text.strip()[:MAX_TEXT_CHARS] if text.strip() else None
    except ImportError:
        print("  [WARN] python-docx not installed — skipping .docx", file=sys.stderr)
    except Exception as e:
        print(f"  [WARN] .docx extraction failed: {e}", file=sys.stderr)
    return None


def _extract_xlsx(filepath):
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filepath, read_only=True)
        lines = []
        for ws in wb.worksheets:
            lines.append(f"[Sheet: {ws.title}]")
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    lines.append(" | ".join(cells))
        wb.close()
        text = "\n".join(lines)
        return text.strip()[:MAX_TEXT_CHARS] if text.strip() else None
    except ImportError:
        print("  [WARN] openpyxl not installed — skipping .xlsx", file=sys.stderr)
    except Exception as e:
        print(f"  [WARN] .xlsx extraction failed: {e}", file=sys.stderr)
    return None


def _extract_plaintext(filepath):
    try:
        with open(filepath, "r", errors="replace") as f:
            text = f.read(MAX_TEXT_CHARS)
        return text.strip() if text.strip() else None
    except Exception:
        return None


# --- Helpers ---

def format_size(size_bytes):
    if size_bytes < 1024:
        return f"{size_bytes} B"
    if size_bytes < 1024 * 1024:
        return f"{size_bytes // 1024} KB"
    return f"{size_bytes / (1024 * 1024):.1f} MB"


def get_header(msg, name):
    """Extract a header value from gog's message JSON (handles dict or list formats)."""
    headers = msg.get("headers", {})
    if isinstance(headers, dict):
        val = headers.get(name.lower()) or headers.get(name)
        if val:
            return val

    payload = msg.get("payload", msg)
    header_list = payload.get("headers", [])
    if isinstance(header_list, list):
        for h in header_list:
            if h.get("name", "").lower() == name.lower():
                return h.get("value", "")

    return ""


def parse_date_header(date_str):
    """Try to parse an email Date header into WITA-local string."""
    if not date_str:
        return ""
    for fmt in [
        "%a, %d %b %Y %H:%M:%S %z",
        "%d %b %Y %H:%M:%S %z",
        "%a, %d %b %Y %H:%M:%S %Z",
    ]:
        try:
            dt = datetime.strptime(date_str.strip(), fmt)
            return dt.astimezone(WITA).strftime("%Y-%m-%d %H:%M")
        except ValueError:
            continue
    return date_str


def collect_message_ids(results):
    """Extract unique message IDs from gog search results."""
    ids = set()
    for item in results:
        if isinstance(item, str):
            ids.add(item)
        elif isinstance(item, dict):
            if "id" in item:
                ids.add(item["id"])
            for m in item.get("messages", []):
                if isinstance(m, dict) and "id" in m:
                    ids.add(m["id"])
    return sorted(ids)


# --- Main ---

def main():
    parser = argparse.ArgumentParser(
        description="Scan daily emails, extract body text and attachments to memory file",
    )
    parser.add_argument("--date", help="Target date YYYY-MM-DD (default: yesterday WITA)")
    parser.add_argument("--dry-run", action="store_true", help="Print output, don't write files")
    args = parser.parse_args()

    if args.date:
        target_date = args.date
    else:
        yesterday = datetime.now(WITA) - timedelta(days=1)
        target_date = yesterday.strftime("%Y-%m-%d")

    print(f"Scanning emails for {target_date}...")

    # Idempotency check
    memory_file = os.path.join(MEMORY_DIR, f"{target_date}.md")
    if not args.dry_run and os.path.exists(memory_file):
        with open(memory_file, "r") as f:
            if IDEMPOTENCY_HEADER in f.read():
                print(f"  Already processed ({IDEMPOTENCY_HEADER} exists). Skipping.")
                return

    # Search all emails for the date
    results = search_emails(target_date)
    if not results:
        print("  No emails found.")
        return

    msg_ids = collect_message_ids(results)
    if not msg_ids:
        print("  No message IDs found in search results.")
        return

    print(f"  Found {len(msg_ids)} message(s) to process.")

    output_sections = []
    temp_files = []

    for msg_id in msg_ids:
        msg_data = get_message(msg_id)
        if not msg_data:
            continue

        msg = msg_data.get("message", msg_data)
        payload = msg.get("payload", msg)

        sender = get_header(msg, "From") or "Unknown"
        subject = get_header(msg, "Subject") or "(no subject)"
        received = parse_date_header(get_header(msg, "Date"))

        section_lines = [
            f"#### From: {sender} — {subject}",
            f"_Received: {received}_",
            "",
        ]

        # --- Email body text ---
        body_text = find_body_text(payload)
        if body_text:
            truncated = body_text.strip()[:MAX_TEXT_CHARS]
            quoted = "\n".join(f"> {line}" for line in truncated.split("\n"))
            section_lines.append(quoted)
            section_lines.append("")

        # --- Attachments ---
        attachments = find_attachments(payload)
        for att in attachments:
            filename = att.get("filename", "unknown")
            att_id = att.get("body", {}).get("attachmentId", "")
            size = att.get("body", {}).get("size", 0)

            print(f"  Downloading: {filename} ({format_size(size)})...")

            filepath = download_attachment(msg_id, att_id, filename)
            text = None
            if filepath:
                temp_files.append(filepath)
                text = extract_text(filepath, filename)

            section_lines.append(f"**{filename}** ({format_size(size)}):")
            if text:
                quoted = "\n".join(f"> {line}" for line in text.split("\n"))
                section_lines.append(quoted)
            else:
                ext = os.path.splitext(filename.lower())[1]
                if ext in EXTRACTABLE_EXTENSIONS:
                    section_lines.append("> [Text extraction failed]")
                else:
                    section_lines.append("> [Binary/unsupported format — no text extracted]")
            section_lines.append("")

        output_sections.append("\n".join(section_lines))

    # Cleanup temp files
    for fp in temp_files:
        try:
            os.remove(fp)
        except OSError:
            pass

    if not output_sections:
        print("  No email content extracted.")
        return

    output = IDEMPOTENCY_HEADER + "\n\n" + "\n".join(output_sections)

    if args.dry_run:
        print("\n--- DRY RUN OUTPUT ---\n")
        print(output)
        print("\n--- END DRY RUN ---")
    else:
        os.makedirs(MEMORY_DIR, exist_ok=True)
        with open(memory_file, "a") as f:
            f.write("\n\n" + output)
        print(f"  Appended to {memory_file}")

    print("Done.")


if __name__ == "__main__":
    main()
